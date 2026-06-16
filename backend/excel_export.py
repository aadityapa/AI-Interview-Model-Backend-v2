"""Build .xlsx evaluation workbooks from stored HR records."""

from __future__ import annotations

import json
from io import BytesIO
from typing import Any, Dict


def build_evaluation_xlsx(record: Dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws0 = wb.active
    ws0.title = "Summary"
    rep = record.get("report") or {}
    ws0.append(["Field", "Value"])
    ws0["A1"].font = bold
    ws0["B1"].font = bold
    rows = [
        ("Interview ID", str(record.get("id", ""))),
        ("Candidate", str(record.get("candidate_name", ""))),
        ("Email", str(record.get("candidate_email", ""))),
        ("Created", str(record.get("created_at", ""))),
        ("Created Date (IST)", str(record.get("created_date_ist", ""))),
        ("Created Time (IST)", str(record.get("created_time_ist", ""))),
        ("Updated", str(record.get("updated_at", ""))),
        ("Updated Date (IST)", str(record.get("updated_date_ist", ""))),
        ("Updated Time (IST)", str(record.get("updated_time_ist", ""))),
        ("Difficulty", str(record.get("difficulty", ""))),
        ("Model", str(record.get("model", ""))),
        ("Final skills", ", ".join(record.get("skills") or [])),
        ("Overall score", rep.get("overall_score", "")),
        ("Overall fitment", rep.get("overall_fitment", "")),
        ("Recommendation", rep.get("recommendation", "")),
        ("Summary", rep.get("summary", "")),
    ]
    for k, v in rows:
        ws0.append([k, v])
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 90
    for row in ws0.iter_rows(min_row=2, max_row=ws0.max_row, min_col=2, max_col=2):
        for c in row:
            c.alignment = wrap

    ws1 = wb.create_sheet("Skill scores")
    ws1.append(["Skill", "Score /10", "Evidence"])
    for c in ws1[1]:
        c.font = bold
    for item in rep.get("skill_scores") or []:
        if not isinstance(item, dict):
            continue
        ws1.append(
            [
                str(item.get("skill", "")),
                item.get("score", ""),
                str(item.get("evidence", "")),
            ]
        )
    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 80
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=3, max_col=3):
        for c in row:
            c.alignment = wrap

    ws2 = wb.create_sheet("Strengths & gaps")
    ws2.append(["Type", "Detail"])
    for c in ws2[1]:
        c.font = bold
    for s in rep.get("strengths") or []:
        ws2.append(["Strength", str(s)])
    for g in rep.get("gaps") or []:
        ws2.append(["Gap", str(g)])
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 90

    ws3 = wb.create_sheet("Q and A")
    ws3.append(["Candidate", str(record.get("candidate_name", ""))])
    ws3.append(["Interview Date (IST)", str(record.get("created_date_ist", ""))])
    ws3.append(["Interview Time (IST)", str(record.get("created_time_ist", ""))])
    ws3.append([])
    ws3.append(["#", "Question", "Answer"])
    for c in ws3[5]:
        c.font = bold
    qs = record.get("questions") or []
    ans = record.get("answers") or []
    for i, q in enumerate(qs):
        a = ans[i] if i < len(ans) else ""
        ws3.append([i + 1, str(q), str(a)])
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 70
    ws3.column_dimensions["C"].width = 70
    for row in ws3.iter_rows(min_row=6, max_row=ws3.max_row, min_col=2, max_col=3):
        for c in row:
            c.alignment = wrap

    ws4 = wb.create_sheet("Report JSON")
    ws4.append(["Structured report (JSON)"])
    ws4["A1"].font = bold
    ws4.append([json.dumps(rep, indent=2, ensure_ascii=False)])
    ws4.column_dimensions["A"].width = 120
    ws4["A2"].alignment = wrap

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
